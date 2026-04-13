"""키워드 분석 API 라우터"""
import asyncio
import json
import io
import csv
import os
from datetime import datetime
from urllib.parse import quote

import openpyxl
import requests as req
from fastapi import APIRouter, Request, UploadFile, File
from fastapi.responses import Response
from src.services.sse_helper import sse_dict, SSEResponse

from src.services.config import (
    executor, selenium_semaphore, GEMINI_API_KEY, PROGRESS_FILE, EXPAND_PROGRESS_FILE, KEYWORD_DB_ID,
)
from src.services.common import error_response, valid_kw
from src.services.naver_search import (
    autocomplete, expand_selenium, search_volume, analyze_serp, ad_signature,
)
from src.services.notion_client import (
    save_keyword_to_notion, save_progress, load_progress,
    save_expand_progress, load_expand_progress, clear_expand_progress,
)
from src.services.selenium_pool import create_driver

router = APIRouter()


# ── 엑셀/CSV 업로드 ──

@router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    content = await file.read()
    keywords = []
    if file.filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.reader(io.StringIO(text))
        for row in reader:
            if row and row[0].strip():
                keywords.append(row[0].strip())
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] and str(row[0]).strip():
                keywords.append(str(row[0]).strip())
    return {'keywords': keywords}


# ── 키워드 확장 (SSE) ──

@router.post("/expand")
async def expand_keywords(request: Request):
    body = await request.json()
    seeds = body.get('keywords', [])
    mode = body.get('mode', 'excel')  # 'excel' (1회 확장) 또는 'repeat' (반복 확장)
    resume = body.get('resume', False)

    async def generate():
      all_kws = {}
      try:
        visited = set()
        start_index = 0
        queue = []
        round_num = 0
        resumed = False

        loop = asyncio.get_running_loop()

        # ── 이전 진행 복구 ──
        if resume:
            prog = load_expand_progress()
            if prog:
                all_kws = prog.get('all_kws', {})
                visited = set(prog.get('visited', []))
                remaining_seeds = prog.get('remaining_seeds', [])
                saved_mode = prog.get('mode', mode)
                queue = prog.get('queue', [])
                round_num = prog.get('round_num', 0)
                resumed = True
                yield sse_dict({'type':'resume','count':len(all_kws),'msg':f'이전 진행 복구: {len(all_kws)}개 키워드 로드'})
                # 복구된 키워드를 프론트에 전달
                for k, src in all_kws.items():
                    yield sse_dict({'type':'keyword','keyword':k,'source':src})
                # resume 시 remaining_seeds를 seeds로 사용
                if saved_mode == 'excel':
                    seeds_to_process = remaining_seeds
                    start_index = len(seeds) - len(remaining_seeds) if remaining_seeds else len(seeds)
                else:
                    seeds_to_process = []  # repeat 모드는 queue로 이어감
            else:
                yield sse_dict({'type':'progress','msg':'복구할 진행 데이터 없음. 처음부터 시작.','cur':0,'total':0})
                seeds_to_process = seeds
        else:
            seeds_to_process = seeds

        # 시드 등록 (resume이 아닐 때만)
        if not resumed:
            for s in seeds:
                all_kws[s] = '시드'
                yield sse_dict({'type':'keyword','keyword':s,'source':'시드'})

        yield sse_dict({'type':'progress','msg':'브라우저 시작 중...','cur':0,'total':0})
        await selenium_semaphore.acquire()
        driver = None
        try:
            driver = await loop.run_in_executor(executor, create_driver)

            if mode == 'excel':
                # ── 모드1: 엑셀 기반 1회 확장 ──
                total = len(seeds)  # 전체 시드 수 (프로그레스 표시용)
                kws_to_expand = seeds_to_process if resume else seeds
                yield sse_dict({'type':'progress','msg':f'엑셀 기반 확장: {len(kws_to_expand)}개 키워드의 자완/연관/함께찾는 수집','cur':start_index,'total':total})

                for i, kw in enumerate(kws_to_expand):
                    cur_idx = start_index + i + 1
                    if kw in visited:
                        continue
                    visited.add(kw)

                    yield sse_dict({'type':'progress','msg':f'확장 중: {kw} ({cur_idx}/{total}) | 누적 {len(all_kws)}개','cur':cur_idx,'total':total})

                    # ── 자동완성 (개별 에러 격리) ──
                    try:
                        ac = await loop.run_in_executor(executor, autocomplete, kw)
                        for ak in ac:
                            if valid_kw(ak) and ak not in all_kws:
                                all_kws[ak] = '자동완성'
                                yield sse_dict({'type':'keyword','keyword':ak,'source':'자동완성'})
                    except Exception as ac_err:
                        print(f"[expand] 자동완성 실패 '{kw}': {ac_err}")
                        yield sse_dict({'type':'warning','msg':f'자동완성 실패: {kw} (건너뜀)'})

                    # ── 셀레니움 확장 (개별 에러 격리) ──
                    try:
                        related = await loop.run_in_executor(executor, expand_selenium, driver, kw)
                        for rk, src in related.items():
                            if rk not in all_kws:
                                all_kws[rk] = src
                                yield sse_dict({'type':'keyword','keyword':rk,'source':src})
                    except Exception as sel_err:
                        print(f"[expand] 셀레니움 확장 실패 '{kw}': {sel_err}")
                        yield sse_dict({'type':'warning','msg':f'연관검색어 수집 실패: {kw} (건너뜀)'})

                    # 10개마다 중간 저장
                    if cur_idx % 10 == 0:
                        remaining = kws_to_expand[i+1:]
                        await loop.run_in_executor(
                            executor, save_expand_progress,
                            all_kws, visited, remaining, mode
                        )
                        yield sse_dict({'type':'saved','count':len(all_kws),'msg':f'중간 저장 완료 ({len(all_kws)}개)'})

                    await asyncio.sleep(1.5)

                yield sse_dict({'type':'progress','msg':f'엑셀 기반 확장 완료. 총 {len(all_kws)}개','cur':total,'total':total})

            else:
                # ── 모드2: 반복 확장 (BFS) ──
                if not queue and not resumed:
                    queue = list(seeds)

                while queue:
                    round_num += 1
                    current_batch = list(queue)
                    queue = []

                    yield sse_dict({'type':'progress','msg':f'[라운드 {round_num}] {len(current_batch)}개 키워드 확장 시작 (누적 {len(all_kws)}개)','cur':0,'total':len(current_batch)})

                    for i, kw in enumerate(current_batch):
                        if kw in visited:
                            continue
                        visited.add(kw)

                        yield sse_dict({'type':'progress','msg':f'[라운드 {round_num}] 확장 중: {kw} ({i+1}/{len(current_batch)}) | 누적 {len(all_kws)}개','cur':i+1,'total':len(current_batch)})

                        # ── 자동완성 (개별 에러 격리) ──
                        try:
                            ac = await loop.run_in_executor(executor, autocomplete, kw)
                            for ak in ac:
                                if valid_kw(ak) and ak not in all_kws:
                                    all_kws[ak] = '자동완성'
                                    if ak not in visited:
                                        queue.append(ak)
                                    yield sse_dict({'type':'keyword','keyword':ak,'source':'자동완성'})
                        except Exception as ac_err:
                            print(f"[expand] 자동완성 실패 '{kw}': {ac_err}")
                            yield sse_dict({'type':'warning','msg':f'자동완성 실패: {kw} (건너뜀)'})

                        # ── 셀레니움 확장 (개별 에러 격리) ──
                        try:
                            related = await loop.run_in_executor(executor, expand_selenium, driver, kw)
                            for rk, src in related.items():
                                if rk not in all_kws:
                                    all_kws[rk] = src
                                    if rk not in visited:
                                        queue.append(rk)
                                    yield sse_dict({'type':'keyword','keyword':rk,'source':src})
                        except Exception as sel_err:
                            print(f"[expand] 셀레니움 확장 실패 '{kw}': {sel_err}")
                            yield sse_dict({'type':'warning','msg':f'연관검색어 수집 실패: {kw} (건너뜀)'})

                        await asyncio.sleep(1.5)

                    # 라운드 끝마다 중간 저장
                    await loop.run_in_executor(
                        executor, save_expand_progress,
                        all_kws, visited, [], mode, queue, round_num
                    )
                    yield sse_dict({'type':'saved','count':len(all_kws),'msg':f'라운드 {round_num} 중간 저장 ({len(all_kws)}개)'})

                    new_count = len(queue)
                    yield sse_dict({'type':'progress','msg':f'[라운드 {round_num} 완료] 신규 {new_count}개 발견 → 누적 {len(all_kws)}개','cur':len(current_batch),'total':len(current_batch)})

                    if not queue:
                        yield sse_dict({'type':'progress','msg':f'더 이상 새로운 키워드 없음. 총 {round_num}라운드 완료.','cur':1,'total':1})
        finally:
            if driver:
                try:
                    await loop.run_in_executor(executor, driver.quit)
                except Exception as drv_err:
                    print(f"[expand] driver.quit 실패: {drv_err}")
                    try:
                        await loop.run_in_executor(executor, driver.close)
                    except Exception:
                        pass
            selenium_semaphore.release()

        # ── 검색량 조회 (배치별 프로그레스 + 에러 격리) ──
        kw_list = [k for k in all_kws.keys() if valid_kw(k)]
        vol = {}
        batch_size = 5  # search_volume 내부 배치 사이즈와 동일
        total_batches = (len(kw_list) + batch_size - 1) // batch_size if kw_list else 0

        yield sse_dict({'type':'progress','msg':f'검색량 조회 시작 ({len(kw_list)}개, {total_batches}배치)','cur':0,'total':total_batches,'phase':'volume'})

        for batch_idx in range(0, len(kw_list), batch_size):
            batch = kw_list[batch_idx:batch_idx + batch_size]
            cur_batch_num = batch_idx // batch_size + 1
            try:
                batch_vol = await loop.run_in_executor(executor, search_volume, batch)
                vol.update(batch_vol)
            except Exception as vol_err:
                print(f"[expand] 검색량 배치 {cur_batch_num} 실패: {vol_err}")
                yield sse_dict({'type':'warning','msg':f'검색량 조회 실패 (배치 {cur_batch_num}/{total_batches}) — 해당 키워드 0으로 처리'})
                for k in batch:
                    vol[k] = {'pc': 0, 'mo': 0}

            yield sse_dict({'type':'progress','msg':f'검색량 조회 중... ({cur_batch_num}/{total_batches})','cur':cur_batch_num,'total':total_batches,'phase':'volume'})

        # ── 결과 조립 ──
        result_list = []
        for k in kw_list:
            v = vol.get(k, {'pc': 0, 'mo': 0})
            pc = v['pc'] if isinstance(v['pc'], int) else 0
            mo = v['mo'] if isinstance(v['mo'], int) else 0
            result_list.append({
                'keyword': k,
                'source': all_kws[k],
                'pc': pc,
                'mo': mo,
                'total': pc + mo,
            })

        # 완료 → 진행 파일 삭제
        await loop.run_in_executor(executor, clear_expand_progress)

        yield sse_dict({'type':'complete','keywords':result_list,'total':len(result_list)})
      except Exception as e:
        # ── 치명적 에러 시에도 지금까지 수집한 키워드 반환 ──
        print(f"[expand_keywords] 치명적 에러: {e}")
        if all_kws:
            kw_list = [k for k in all_kws.keys() if valid_kw(k)]
            partial_list = [{'keyword': k, 'source': all_kws[k], 'pc': 0, 'mo': 0, 'total': 0} for k in kw_list]
            yield sse_dict({'type':'partial_complete','keywords':partial_list,'total':len(partial_list),'error':str(e),'msg':f'오류 발생 — {len(partial_list)}개 키워드 부분 반환 (검색량 미조회)'})
        else:
            yield sse_dict({'type':'error','message':f'키워드 확장 중 오류: {e}'})

    return SSEResponse(generate())


# ── 확장 진행 상태 확인 ──

@router.get("/check-expand-progress")
async def check_expand_progress():
    prog = load_expand_progress()
    if prog:
        return {
            'has_progress': True,
            'count': len(prog.get('all_kws', {})),
            'mode': prog.get('mode', ''),
            'timestamp': prog.get('ts', ''),
        }
    return {'has_progress': False}


# ── 검색량 조회 ──

@router.post("/search-volume")
async def search_volume_api(request: Request):
    """중지 후 검색량만 별도 조회"""
    body = await request.json()
    keywords = body.get('keywords', [])
    loop = asyncio.get_running_loop()
    vol = await loop.run_in_executor(executor, search_volume, keywords)
    return vol


# ── 키워드 분석 (SSE) ──

@router.post("/analyze")
async def analyze_keywords(request: Request):
    body = await request.json()
    keywords = body.get('keywords', [])
    resume = body.get('resume', False)

    async def generate():
      try:
        loop = asyncio.get_running_loop()
        today = datetime.now()
        results = []
        start_idx = 0

        # 이전 진행분 복구
        if resume:
            prog = load_progress()
            if prog:
                results = prog.get('results', [])
                remaining = prog.get('remaining', [])
                keywords_to_process = remaining
                start_idx = len(results)
                yield sse_dict({'type':'resume','existing':results,'start':start_idx})
            else:
                keywords_to_process = keywords
        else:
            keywords_to_process = keywords

        total = start_idx + len(keywords_to_process)

        # 검색량 배치 조회
        yield sse_dict({'type':'progress','msg':'검색량 조회 중...','cur':0,'total':total})
        vol = await loop.run_in_executor(executor, search_volume, keywords_to_process)

        for i, kw in enumerate(keywords_to_process):
            idx = start_idx + i + 1
            yield sse_dict({'type':'progress','msg':f'분석 중: {kw}','cur':idx,'total':total})

            try:
                # SERP 분석
                serp = await loop.run_in_executor(executor, analyze_serp, kw, today)
                v = vol.get(kw, {'pc':0,'mo':0})
            except Exception as kw_err:
                print(f"[analyze] 키워드 '{kw}' 분석 실패: {kw_err}")
                serp = {'competition': '-', 'content_tab_rank': '-', 'content_format': '-', 'articles': [], 'top6_tabs': '-'}
                v = {'pc': 0, 'mo': 0}

            row = {
                'keyword': kw,
                'pc': v['pc'], 'mo': v['mo'], 'total': v['pc'] + v['mo'],
                'competition': serp.get('competition', '-'),
                'content_tab_rank': serp.get('content_tab_rank', '-'),
                'content_format': serp.get('content_format', '-'),
                'contact_point': '',
                'articles': serp.get('articles', []),
                'top6_tabs': serp.get('top6_tabs', '-'),
            }
            results.append(row)

            yield sse_dict({'type':'result','row':row,'cur':idx,'total':total})

            # 30개마다 중간저장
            if idx % 30 == 0:
                remaining = keywords_to_process[i+1:]
                await loop.run_in_executor(executor, save_progress, results, remaining)
                yield sse_dict({'type':'saved','count':len(results)})

            await asyncio.sleep(1.5)

        # 완료 시 진행파일 삭제
        if os.path.exists(PROGRESS_FILE):
            os.remove(PROGRESS_FILE)

        yield sse_dict({'type':'complete','total':len(results)})
      except Exception as e:
        print(f"[analyze_keywords] 에러: {e}")
        yield sse_dict({'type':'error','message':f'키워드 분석 중 오류: {e}'})

    return SSEResponse(generate())


# ── 노션 저장 ──

@router.post("/save-notion")
async def save_notion(request: Request):
    body = await request.json()
    items = body.get('items', [])
    loop = asyncio.get_running_loop()
    success = 0
    for item in items:
        ok = await loop.run_in_executor(executor, save_keyword_to_notion, item)
        if ok:
            success += 1
        await asyncio.sleep(0.3)
    return {'success': success, 'total': len(items)}


# ── 진행 상태 확인 ──

@router.get("/check-progress")
async def check_progress():
    prog = load_progress()
    if prog:
        return {'has_progress': True, 'count': len(prog.get('results', [])), 'timestamp': prog.get('ts', '')}
    return {'has_progress': False}


# ── Gemini Flash API (접촉지점 판별) ──

GEMINI_SYSTEM_PROMPT = '아래 키워드 각각을 검색하는 사람이 구매여정 6단계 중 어디에 있는지 판별해. 키워드 자체의 맥락을 보고 판단해. 0_무지: 문제 자체를 모름 1_인지: 문제 인식 (예: 전립선 증상) 2_호기심: 해결책 탐색 (예: 전립선에 좋은 음식) 3_정보습득: 해결책 찾는 중 (예: 전립선 영양제) 4_의심: 비교/검토 (예: 전립선 영양제 부작용) 5_구매직전: 구매 의사 명확 (예: 전립선 영양제 추천) JSON으로 응답해: {"키워드": "단계", ...}'


def _call_gemini_contact_point(keywords_batch):
    """Gemini Flash로 접촉지점 판별 (50개씩)"""
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key=" + GEMINI_API_KEY
    user_msg = '\n'.join(keywords_batch)
    payload = {
        'system_instruction': {'parts': [{'text': GEMINI_SYSTEM_PROMPT}]},
        'contents': [{'parts': [{'text': user_msg}]}],
        'generationConfig': {'temperature': 0.1, 'responseMimeType': 'application/json'},
    }
    try:
        r = req.post(url, json=payload, timeout=30)
        if r.status_code == 200:
            data = r.json()
            text = data['candidates'][0]['content']['parts'][0]['text']
            return json.loads(text)
        else:
            print("Gemini API error %d: %s" % (r.status_code, r.text[:200]))
            return {}
    except Exception as e:
        print("Gemini API error: %s" % e)
        return {}


@router.post("/contact-point")
async def contact_point(request: Request):
    """키워드 접촉지점 판별 (Gemini Flash, 50개씩 배치)"""
    body = await request.json()
    keywords = body.get('keywords', [])
    loop = asyncio.get_running_loop()
    result = {}
    batch_size = 50
    for i in range(0, len(keywords), batch_size):
        batch = keywords[i:i+batch_size]
        batch_result = await loop.run_in_executor(executor, _call_gemini_contact_point, batch)
        result.update(batch_result)
    return result


# ── 엑셀 다운로드 ──

@router.post("/download-excel")
async def download_excel(request: Request):
    """서버 사이드 엑셀 생성 + 다운로드"""
    body = await request.json()
    items = body.get('items', [])
    filename = body.get('filename', '키워드')
    sheet_name = body.get('sheet', 'Sheet1')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if items:
        # 헤더
        headers = list(items[0].keys())
        ws.append(headers)
        # 데이터
        for item in items:
            ws.append([item.get(h, '') for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_filename = quote(f"{filename}.xlsx")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}",
        }
    )
